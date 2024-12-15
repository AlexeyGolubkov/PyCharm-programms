# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
from collections import defaultdict
from openpyxl import load_workbook
from utils_for_print_output import *
from utils_operation_net import *

def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi('Start')

file_path = 'C:\\data\\'
file_input = 'Tab_example.xlsx'
file_out = 'output.txt' #there is temp stub/it will name of column in end

# Загружаем Excel-файл
workbook = load_workbook(filename=file_path + file_input)
sheet = workbook.active
row_count = sheet.max_row
column_count = sheet.max_column
# Создаем пустой словарь
excel_map_apn = defaultdict(list)
aggregated_array_network = {}
remaining_array_networks = {}
for k in range(2, column_count + 1, 1):
    array_network_for_all_apn = []
    for i in range(2, row_count + 1, 1):
        array_network = []
        cell_value = sheet.cell(row=i, column=k).value
        key = sheet.cell(row=i, column=1).value

        if '\t' in cell_value or ' ' in cell_value or '\n' in cell_value:
            array_network = cell_value.split()

        else:
            array_network =[]
            array_network.append(cell_value)

        excel_map_apn[key].extend(array_network)



        # array_network_for_all_apn = array_network_for_all_apn.append(array_network)

        print('excel_map_apn:', excel_map_apn)
print_in_output_file_name_from_line1_current_column(file_path + str(sheet.cell(row=1, column=k).value+'.txt'),
                                                            excel_map_apn)

