# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

from openpyxl import load_workbook


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi('Start')

file_path = 'C:\\data\\'
file_input = 'Tab_example.xlsx'
file_out = 'output.txt'

# Загружаем Excel-файл
workbook = load_workbook(filename=file_path + file_input)
sheet = workbook.active

# Создаем пустой словарь
excel_map = {}

# Проходим по каждой колонке и создаем соответствующие ключи и значения
for col_idx in range(1, sheet.max_column + 1):
    key = sheet.cell(row=1, column=col_idx).value
    values = []

    for row_idx in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row_idx, column=col_idx).value
        if value is not None:
            values.append(value)

    excel_map[key] = values

    # Вывод результата
    print(excel_map)
# Найдем индекс колонки с названием "Регион"
region_col_index = None
for i in range(1, sheet.max_column + 1):
    cell_value = sheet.cell(row=1, column=i).value
    if cell_value == "Region":
        region_col_index = i
        break

# Проверка, существует ли такая колонка
if region_col_index is None:
    raise ValueError("Колонка 'Регион' не найдена.")

# Распечатаем шестую строку из колонки "Регион"
sixth_row_region = sheet.cell(row=6, column=region_col_index).value
print(sixth_row_region)

# Получаем значение из второй колонки третьей строки
cell_value = sheet.cell(row=1, column=1).value
print(f'Значение из 1 колонки & 1 строки: {cell_value}')

# Находим строку с заголовком 'Region' и получаем значение из пятой строки этого столбца
for row in sheet.iter_rows():
    for cell in row:
        if cell.value == 'Region':
            region_column_index = cell.column
            break
    else:
        continue
    break

region_cell = sheet.cell(row=5, column=region_column_index).value
print(f'Значение из пятой строки столбца "Region": {region_cell}')

from collections import defaultdict

# Создаем словарь с заводским значением по умолчанию - пустым списком
my_dict = defaultdict(list)

# Добавляем данные без проверки наличия ключа
my_dict['key1'].append('apple')
my_dict['key2'].append('banana')

# Проверка содержимого словаря
print(dict(my_dict))  # Преобразование в обычный словарь для вывода
# Вывод: {'key1': ['apple'], 'key2': ['banana']}

import ipaddress

# Исходная сеть
network = ipaddress.IPv4Network('192.168.0.0/24')

# Разделение сети на две равные части
new_prefix = network.prefixlen + 1
subnets = list(network.subnets(new_prefix=new_prefix))

# Выводим новые сети
for subnet in subnets:
    print(subnet)


def divide(a, b):
    # Объявляем переменную result вне блока try/except,
    # чтобы она была доступна после завершения блока
    result = None

    try:
        # Попытка выполнения операции деления
        result = a / b

        # Можно добавить несколько строк кода внутри try-блока
        print(f"{a} делить на {b} равно {result}")

    except ZeroDivisionError as e:
        # Обработка исключения деления на ноль
        print("Ошибка: деление на ноль!")
        print(e)

    except TypeError as e:
        # Обработка исключения типа данных
        print("Ошибка: неверный тип данных!")
        print(e)

    except Exception as e:
        # Обработка всех остальных возможных исключений
        print("Произошла неизвестная ошибка:")
        print(e)

    else:
        # Этот блок выполняется, если в try-блоке не возникло никаких исключений
        print("Операция прошла успешно.")

    finally:
        # Этот блок всегда выполняется независимо от того, было исключение или нет
        print("Блок finally выполнен.")

    return result


# Вызываем функцию с различными аргументами
print(divide(10, 0))  # Ошибка деления на ноль
print(divide(20, 5))  # Успешное выполнение
print(divide('abc', 4))  # Ошибка типа данных


def write_to_file(dictionary, filename):
    # Открываем файл для записи
    with open(filename, 'w') as file:
        # Сортируем ключи словаря по алфавиту
        for key in sorted(dictionary.keys()):
            # Получаем значение по ключу
            value = dictionary[key]

            # Фильтруем строковые элементы из списка
            string_elements = [item for item in value if isinstance(item, str)]

            # Сортируем строковые элементы по алфавиту
            string_elements.sort()

            # Записываем строку в файл
            for element in string_elements:
                file.write(f"{element}\n")


# Пример словаря
data = {
    "key1": ["apple", 123, "banana"],
    "key2": ["orange", "kiwi", 456],
    "key3": ["pear", "mango", 789]
}

# Имя файла для вывода
output_filename = "output.txt"

# Вызов функции для записи данных в файл
write_to_file(data, output_filename)
